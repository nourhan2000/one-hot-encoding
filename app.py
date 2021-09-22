import nltk
from nltk.corpus import stopwords
from openpyxl import load_workbook
from nltk.util import ngrams

binRep = []
book_name = []


def iterSheet(words):
    workbook = load_workbook(filename="books.xlsx")
    sheet = workbook.active
    for value in sheet.iter_rows(min_row=1, values_only=True):
        for x in words:
            if x in value:
                if value[3] not in binRep:
                    binRep.append(value[3])
                    book_name.append(value[4])
    return



def getBooks(msg):
    model = ['want', 'need', 'please', 'plz', 'thanks', 'advance', 'hope', 'find', 'read', 'know', 'well', 'shall',
         'book', 'books','','good','recommend']
    answer= " the books' names are:  "
    tokens = [w.lower() for w in msg]
    words = [word for word in tokens if word.isalpha()]
    stop_words = set(stopwords.words('english'))
    words = [w for w in words if not w in stop_words]
    words = [w for w in words if not w in model]
    iterSheet(words)
    words2 = []

    bigrams = ngrams(words, 2)

    for item in bigrams:
        words2.append(item[0] + ' ' + item[1])

    iterSheet(words2)
    
    for b in book_name:
        answer = answer +" "+ str(b)
    answer = answer + "\n the binary representation is: "
    for r in binRep:
        answer =  answer + " "+str(r)
    return str(answer)
    


#getBooks('i want art')
print(binRep)
print(book_name)